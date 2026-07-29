from flask import Flask, render_template, request, send_file
import pandas as pd
import numpy as np
import datetime
import os
import time
import re
import io
from html.parser import HTMLParser
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# PDF Generation imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ==========================================
# MASTER CITY-STATE MAPPING DICTIONARY
# ==========================================
CITY_STATE_MAP = {
    'chennai': 'Tamil Nadu',
    'hitech chennai': 'Tamil Nadu',
    'coimbatore': 'Tamil Nadu',
    'madurai': 'Tamil Nadu',
    'pondichery': 'Tamil Nadu',
    'salem': 'Tamil Nadu',
    'trichy': 'Tamil Nadu',
    'chidambaram': 'Tamil Nadu',
    'tirupur': 'Tamil Nadu',
    'kanchipuram': 'Tamil Nadu',
    'erode': 'Tamil Nadu',
    'thiruvallur': 'Tamil Nadu',
    'thiruttani': 'Tamil Nadu',
    'vellore': 'Tamil Nadu',
    'tirunelveli': 'Tamil Nadu',
    'nagercoil': 'Tamil Nadu',
    'villupuram': 'Tamil Nadu',
    'thanjavur': 'Tamil Nadu',
    'hyderabad': 'Telangana',
    'vijayawada': 'Andhra pradesh',
    'vizag': 'Andhra pradesh',
    'guntur': 'Andhra pradesh',
    'alleppey': 'Kerala',
    'calicut': 'Kerala',
    'cochin': 'Kerala',
    'kannur': 'Kerala',
    'karunagappally': 'Kerala',
    'perumbavoor': 'Kerala',
    'thrissur': 'Kerala',
    'trivandrum': 'Kerala',
    'bangalore': 'Karnataka',
    'kolar': 'Karnataka',
    'belgaum': 'Karnataka',
    'mangalore': 'Karnataka',
    'mysore': 'Karnataka',
    'hubli': 'Karnataka',
    'shimoga': 'Karnataka',
    'tumkur': 'Karnataka'
}

# ==========================================
# EXACT HTML TABLE PARSER
# ==========================================
class ExactHTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self.current_row = []
        self.current_cell = []
        self.in_cell = False
        self.in_table = False

    def handle_starttag(self, tag, attrs):
        if tag.lower() == 'table':
            self.in_table = True
        elif self.in_table and tag.lower() == 'tr':
            self.current_row = []
        elif self.in_table and tag.lower() in ['td', 'th']:
            self.in_cell = True
            self.current_cell = []

    def handle_endtag(self, tag):
        if tag.lower() == 'table':
            self.in_table = False
        elif self.in_table and tag.lower() == 'tr':
            if self.current_row:
                self.rows.append(self.current_row)
            self.current_row = []
        elif self.in_table and tag.lower() in ['td', 'th']:
            self.in_cell = False
            cell_text = "".join(self.current_cell).strip()
            self.current_row.append(cell_text)
            self.current_cell = []

    def handle_data(self, data):
        if self.in_cell:
            self.current_cell.append(data)

def parse_html_exact_strings(file_path):
    for enc in ['utf-8', 'latin1', 'cp1252', 'utf-16', 'utf-16-le']:
        try:
            with open(file_path, 'r', encoding=enc, errors='ignore') as f:
                content = f.read()
            if '<table' in content.lower():
                parser = ExactHTMLTableParser()
                parser.feed(content)
                if parser.rows and len(parser.rows) > 1:
                    headers = parser.rows[0]
                    data = parser.rows[1:]
                    max_cols = max(len(r) for r in parser.rows)
                    headers = headers + [f"Unnamed_{i}" for i in range(len(headers), max_cols)]
                    padded_data = [r + [''] * (max_cols - len(r)) for r in data]
                    df = pd.DataFrame(padded_data, columns=headers)
                    return df
        except Exception:
            pass
    return None

def read_any_excel_file(file_path, sheet_name=0):
    df_html = parse_html_exact_strings(file_path)
    if df_html is not None and not df_html.empty:
        return df_html

    for sep in ['\t', ',', ';', '|']:
        for enc in ['utf-8', 'latin1', 'cp1252', 'utf-16']:
            try:
                df = pd.read_csv(file_path, sep=sep, encoding=enc, low_memory=False, dtype=str, keep_default_na=False)
                if len(df.columns) > 1 and len(df) > 0:
                    return df
            except Exception:
                pass

    for engine_name in ['openpyxl', 'xlrd', 'pyxlsb']:
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name, engine=engine_name)
        except Exception:
            pass

    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        pass

    raise ValueError(f"Could not read uploaded file: {os.path.basename(file_path)}. Please re-save as standard Excel (.xlsx) and try.")

def format_custom_dates(df):
    for col in df.columns:
        col_str = str(col).lower()
        if any(k in col_str for k in ['date', 'time', 'dt', 'pickup', 'drop', 'beat']):
            try:
                s_dt = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
                valid_dt = s_dt.dropna()
                if len(valid_dt) > 0:
                    has_time = (valid_dt.dt.hour != 0).any() or (valid_dt.dt.minute != 0).any() or (valid_dt.dt.second != 0).any()
                    if has_time:
                        formatted = s_dt.dt.strftime('%d-%m-%Y %H:%M:%S')
                    else:
                        formatted = s_dt.dt.strftime('%d-%m-%Y')
                    df[col] = formatted.fillna(df[col])
            except Exception:
                pass
    return df

def time_to_minutes(val):
    if pd.isna(val): return np.nan
    val_str = str(val).strip()
    if ':' in val_str:
        parts = val_str.split(':')
        try: return int(parts[0]) * 60 + int(parts[1])
        except ValueError: return np.nan
    try: return float(val_str) * 60
    except ValueError: return np.nan

def minutes_to_hhmm(val):
    if pd.isna(val) or val <= 0: return "00:00"
    total_mins = int(round(val))
    return f"{total_mins // 60:02d}:{total_mins % 60:02d}"

def parse_time_value(t):
    if isinstance(t, datetime.time): return t
    if isinstance(t, datetime.datetime): return t.time()
    if pd.isna(t): return None
    t_str = str(t).strip()
    if t_str in ['', '-', 'nan', 'NaN', 'None']: return None
    
    for fmt in ('%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M'):
        try:
            return datetime.datetime.strptime(t_str, fmt).time()
        except ValueError:
            pass
            
    try:
        dt = pd.to_datetime(t_str, errors='coerce')
        if pd.notna(dt):
            return dt.time()
    except Exception:
        pass
    return None

def format_hhmm_string(val):
    t_obj = parse_time_value(val)
    if t_obj is not None:
        return t_obj.strftime('%H:%M')
    return "-" if pd.isna(val) or str(val).strip() in ['', '-', 'nan', 'NaN', 'None', 'NaT'] else str(val).strip()

def time_to_total_minutes(t):
    if t is None: return None
    return t.hour * 60 + t.minute

def find_city_col(df):
    return next((c for c in df.columns if str(c).strip().lower() == 'city'), None)

def normalize_bv_columns(df):
    col_map = {}
    for col in df.columns:
        clean = str(col).replace('[', '').replace(']', '').strip().lower()
        if 'non' in clean and 'adherence' in clean:
            col_map[col] = '[Non Adherence]'
        elif 'adherence' in clean:
            col_map[col] = '[Adherence Visit]'
        elif 'total' in clean and 'visit' in clean:
            col_map[col] = '[Total Visit]'
        elif 'frequency' in clean or ('visit' in clean and 'freq' in clean):
            col_map[col] = '[Visit Frequency]'
    
    df = df.rename(columns=col_map)
    for std_col in ['[Visit Frequency]', '[Total Visit]', '[Adherence Visit]', '[Non Adherence]']:
        if std_col not in df.columns:
            df[std_col] = 0
        else:
            df[std_col] = pd.to_numeric(df[std_col], errors='coerce').fillna(0)
    return df

# ==========================================
# PROJECT 1: BEAT PERFORMANCE LOGIC & TOTAL/AVG ROW
# ==========================================
def calculate_period_totals(df1, df2, df3, prefix=""):
    totals = {}
    if not df1.empty:
        df1 = normalize_bv_columns(df1)
        tot_freq = df1['[Visit Frequency]'].sum()
        tot_visit = df1['[Total Visit]'].sum()
        tot_adh = df1['[Adherence Visit]'].sum()
        tot_non_adh = df1['[Non Adherence]'].sum()
        tot_tracked = tot_adh + tot_non_adh

        v_comp = (tot_visit / tot_freq * 100) if tot_freq > 0 else 0
        on_time = (tot_adh / tot_tracked * 100) if tot_tracked > 0 else 0

        totals[f'{prefix}Visit Compliance'] = f"{round(v_comp, 1)}%"
        totals[f'{prefix}On Time %'] = f"{round(on_time, 1)}%"
    else:
        totals[f'{prefix}Visit Compliance'] = "-"
        totals[f'{prefix}On Time %'] = "-"

    if not df2.empty:
        within_30 = df2['is_within_30m'].sum()
        tot_cnt = df2['total_count'].sum()
        p_30 = (within_30 / tot_cnt * 100) if tot_cnt > 0 else 0
        totals[f'{prefix}30mts %'] = f"{round(p_30, 1)}%"
    else:
        totals[f'{prefix}30mts %'] = "0.0%"

    if not df3.empty:
        df3_city_limit = df3[df3['Radius_Clean'].str.contains('city', na=False)]
        df3_out = df3[df3['Radius_Clean'].str.contains('outskirt', na=False)]
        
        tat_c = df3_city_limit['TAT_Minutes'].mean() if not df3_city_limit.empty else np.nan
        tat_o = df3_out['TAT_Minutes'].mean() if not df3_out.empty else np.nan
        tat_all = df3['TAT_Minutes'].mean() if not df3.empty else np.nan

        totals[f'{prefix}P2D-City'] = minutes_to_hhmm(tat_c)
        totals[f'{prefix}P2D-Outskirt'] = minutes_to_hhmm(tat_o)
        totals[f'{prefix}P2D-Overall'] = minutes_to_hhmm(tat_all)
    else:
        totals[f'{prefix}P2D-City'] = "00:00"
        totals[f'{prefix}P2D-Outskirt'] = "00:00"
        totals[f'{prefix}P2D-Overall'] = "00:00"

    return totals

def calculate_sheet_metrics(df1, df2, df3, group_cols, prefix=""):
    if not df1.empty:
        df1 = normalize_bv_columns(df1)
        f1_grouped = df1.groupby(group_cols).agg({
            '[Visit Frequency]': 'sum', '[Total Visit]': 'sum',
            '[Adherence Visit]': 'sum', '[Non Adherence]': 'sum'
        }).reset_index()
        
        f1_grouped[f'{prefix}Visit Compliance'] = np.where(
            f1_grouped['[Visit Frequency]'] > 0,
            (f1_grouped['[Total Visit]'] / f1_grouped['[Visit Frequency]']) * 100,
            0
        )
        total_tracked = f1_grouped['[Adherence Visit]'] + f1_grouped['[Non Adherence]']
        f1_grouped[f'{prefix}On Time %'] = np.where(
            total_tracked > 0,
            (f1_grouped['[Adherence Visit]'] / total_tracked) * 100,
            0
        )
        f1_grouped[f'{prefix}Visit Compliance'] = f1_grouped[f'{prefix}Visit Compliance'].fillna(0).round(1).astype(str) + '%'
        f1_grouped[f'{prefix}On Time %'] = f1_grouped[f'{prefix}On Time %'].fillna(0).round(1).astype(str) + '%'
        m1 = f1_grouped[group_cols + [f'{prefix}Visit Compliance', f'{prefix}On Time %']]
    else:
        m1 = pd.DataFrame(columns=group_cols + [f'{prefix}Visit Compliance', f'{prefix}On Time %'])

    if not df2.empty:
        f2_grouped = df2.groupby(group_cols).agg({'is_within_30m': 'sum', 'total_count': 'sum'}).reset_index()
        f2_grouped[f'{prefix}30mts %'] = np.where(
            f2_grouped['total_count'] > 0,
            (f2_grouped['is_within_30m'] / f2_grouped['total_count']) * 100,
            0
        )
        f2_grouped[f'{prefix}30mts %'] = f2_grouped[f'{prefix}30mts %'].fillna(0).round(1).astype(str) + '%'
        m2 = f2_grouped[group_cols + [f'{prefix}30mts %']]
    else:
        m2 = pd.DataFrame(columns=group_cols + [f'{prefix}30mts %'])

    if not df3.empty:
        df3_city_limit = df3[df3['Radius_Clean'].str.contains('city', na=False)]
        df3_out = df3[df3['Radius_Clean'].str.contains('outskirt', na=False)]
        tat_city = df3_city_limit.groupby(group_cols)['TAT_Minutes'].mean().reset_index().rename(columns={'TAT_Minutes': f'{prefix}P2D-City'})
        tat_out = df3_out.groupby(group_cols)['TAT_Minutes'].mean().reset_index().rename(columns={'TAT_Minutes': f'{prefix}P2D-Outskirt'})
        tat_overall = df3.groupby(group_cols)['TAT_Minutes'].mean().reset_index().rename(columns={'TAT_Minutes': f'{prefix}P2D-Overall'})
        m3 = pd.merge(tat_city, tat_out, on=group_cols, how='outer')
        m3 = pd.merge(m3, tat_overall, on=group_cols, how='outer')
        for col in [f'{prefix}P2D-City', f'{prefix}P2D-Outskirt', f'{prefix}P2D-Overall']:
            if col in m3.columns: m3[col] = m3[col].apply(minutes_to_hhmm)
            else: m3[col] = "00:00"
    else:
        m3 = pd.DataFrame(columns=group_cols + [f'{prefix}P2D-City', f'{prefix}P2D-Outskirt', f'{prefix}P2D-Overall'])

    return pd.merge(pd.merge(m1, m2, on=group_cols, how='outer'), m3, on=group_cols, how='outer')

def add_total_row_to_df(df, df1_ftd, df2_ftd, df3_ftd, df1_mtd, df2_mtd, df3_mtd, first_col='State'):
    if df.empty:
        return df

    ftd_tot = calculate_period_totals(df1_ftd, df2_ftd, df3_ftd, prefix="FTD ")
    mtd_tot = calculate_period_totals(df1_mtd, df2_mtd, df3_mtd, prefix="MTD ")

    total_row = {col: "" for col in df.columns}
    total_row[first_col] = "Total/Avg"

    for k, v in ftd_tot.items():
        if k in df.columns:
            total_row[k] = v
            
    for k, v in mtd_tot.items():
        if k in df.columns:
            total_row[k] = v

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

def process_hub_data(df1_ftd, df2_ftd, df3_ftd, df1_mtd, df2_mtd, df3_mtd, target_city_name):
    f1_c_ftd = df1_ftd[df1_ftd['City'].astype(str).str.lower() == target_city_name.lower()].copy()
    f2_c_ftd = df2_ftd[df2_ftd['City'].astype(str).str.lower() == target_city_name.lower()].copy()
    f3_c_ftd = df3_ftd[df3_ftd['City'].astype(str).str.lower() == target_city_name.lower()].copy()
    
    f1_c_mtd = df1_mtd[df1_mtd['City'].astype(str).str.lower() == target_city_name.lower()].copy()
    f2_c_mtd = df2_mtd[df2_mtd['City'].astype(str).str.lower() == target_city_name.lower()].copy()
    f3_c_mtd = df3_mtd[df3_mtd['City'].astype(str).str.lower() == target_city_name.lower()].copy()

    for df_item in [f1_c_ftd, f1_c_mtd, f2_c_ftd, f2_c_mtd, f3_c_ftd, f3_c_mtd]:
        h_col = [c for c in df_item.columns if 'hub' in c.lower()]
        if h_col: df_item['Hub'] = df_item[h_col[0]].astype(str).str.strip().str.upper()

    group_cols = ['State', 'City', 'Hub']
    ftd_res = calculate_sheet_metrics(f1_c_ftd, f2_c_ftd, f3_c_ftd, group_cols, prefix="FTD ")
    mtd_res = calculate_sheet_metrics(f1_c_mtd, f2_c_mtd, f3_c_mtd, group_cols, prefix="MTD ")
    
    final_hub = pd.merge(ftd_res, mtd_res, on=group_cols, how='outer')
    for pfx in ["FTD ", "MTD "]:
        if f'{pfx}Visit Compliance' in final_hub.columns: final_hub[f'{pfx}Visit Compliance'] = final_hub[f'{pfx}Visit Compliance'].fillna("-")
        if f'{pfx}On Time %' in final_hub.columns: final_hub[f'{pfx}On Time %'] = final_hub[f'{pfx}On Time %'].fillna("-")
        if f'{pfx}30mts %' in final_hub.columns: final_hub[f'{pfx}30mts %'] = final_hub[f'{pfx}30mts %'].fillna("0.0%")
        for tat_c in [f'{pfx}P2D-City', f'{pfx}P2D-Outskirt', f'{pfx}P2D-Overall']:
            if tat_c in final_hub.columns: final_hub[tat_c] = final_hub[tat_c].fillna("00:00")

    if not final_hub.empty:
        final_hub = add_total_row_to_df(final_hub, f1_c_ftd, f2_c_ftd, f3_c_ftd, f1_c_mtd, f2_c_mtd, f3_c_mtd, first_col='State')
            
    return final_hub

def write_not_dropped_sheet(writer, df3):
    if df3.empty:
        return
    
    status_col = next((c for c in df3.columns if 'lab' in c.lower() and 'status' in c.lower()), None)
    if not status_col:
        status_col = next((c for c in df3.columns if 'status' in c.lower()), None)
        
    if not status_col:
        return
        
    df3_nd = df3[df3[status_col].astype(str).str.strip().str.lower().str.contains('not drop', na=False)].copy()
    if df3_nd.empty:
        return
        
    if 'DT_PARSED' not in df3_nd.columns:
        date_col_f3 = [c for c in df3_nd.columns if 'DATE' in c.upper() or 'PICKUP' in c.upper() or 'TIME' in c.upper()][0]
        df3_nd['DT_PARSED'] = pd.to_datetime(df3_nd[date_col_f3], errors='coerce', dayfirst=True)
        
    if 'City' not in df3_nd.columns:
        c3 = find_city_col(df3_nd)
        if c3: df3_nd['City'] = df3_nd[c3].astype(str).str.strip().str.title()
        else: df3_nd['City'] = ''
        
    if 'Hub' not in df3_nd.columns:
        h_col = [c for c in df3_nd.columns if 'hub' in c.lower()]
        if h_col: df3_nd['Hub'] = df3_nd[h_col[0]].astype(str).str.strip().str.upper()
        else: df3_nd['Hub'] = ''

    df3_nd = df3_nd.dropna(subset=['DT_PARSED'])
    if df3_nd.empty:
        return

    wb = writer.book
    ws = wb.create_sheet('Not Dropped')

    def append_pivot_table(start_row, filter_city=None):
        if filter_city:
            sub = df3_nd[df3_nd['City'].astype(str).str.lower() == filter_city.lower()].copy()
            row_col = 'Hub'
        else:
            sub = df3_nd.copy()
            row_col = 'City'

        if sub.empty:
            return start_row

        sub['DATE_FMT'] = sub['DT_PARSED'].dt.strftime('%d-%b')
        sub_sorted = sub.sort_values('DT_PARSED')
        dates_sorted = list(dict.fromkeys(sub_sorted['DATE_FMT']))

        pivot = pd.crosstab(sub[row_col], sub['DATE_FMT']).reindex(columns=dates_sorted, fill_value=0)
        pivot['Grand Total'] = pivot.sum(axis=1)
        total_row = pivot.sum(axis=0)
        total_row.name = 'Grand Total'
        pivot = pd.concat([pivot, pd.DataFrame(total_row).T])

        curr = start_row
        ws.cell(row=curr, column=1, value='LABDROP STATUS').font = Font(bold=True)
        ws.cell(row=curr, column=2, value='Not Dropped')
        curr += 1

        if filter_city:
            ws.cell(row=curr, column=1, value='CITY').font = Font(bold=True)
            ws.cell(row=curr, column=2, value=filter_city.title())
            curr += 1

        curr += 1

        ws.cell(row=curr, column=1, value='Count of PICKUP').font = Font(bold=True)
        ws.cell(row=curr, column=2, value='Column Labels').font = Font(bold=True)
        curr += 1

        headers = ['Row Labels'] + dates_sorted + ['Grand Total']
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=col_idx, value=h_text)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'))
        curr += 1

        for row_label, row_data in pivot.iterrows():
            is_grand_total = (row_label == 'Grand Total')
            cell_lbl = ws.cell(row=curr, column=1, value=row_label)
            if is_grand_total:
                cell_lbl.font = Font(bold=True)

            for col_idx, val in enumerate(row_data, 2):
                cell_val = ws.cell(row=curr, column=col_idx)
                if val > 0:
                    cell_val.value = int(val)
                else:
                    cell_val.value = ""

                if is_grand_total or col_idx == len(headers):
                    cell_val.font = Font(bold=True)

            if is_grand_total:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=curr, column=c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            curr += 1

        return curr + 2

    r = append_pivot_table(1, filter_city=None)
    r = append_pivot_table(r, filter_city='Chennai')
    r = append_pivot_table(r, filter_city='Hitech Chennai')

def process_beat_performance(master_excel_path, ftd_date_raw, mtd_date_raw):
    try:
        df1 = read_any_excel_file(master_excel_path, sheet_name='BV')
        df2 = read_any_excel_file(master_excel_path, sheet_name='EV')
        df3 = read_any_excel_file(master_excel_path, sheet_name='TAT')
    except Exception:
        df1 = read_any_excel_file(master_excel_path, sheet_name=0)
        df2 = pd.DataFrame()
        df3 = pd.DataFrame()

    df1.columns = df1.columns.astype(str).str.strip()
    df1 = normalize_bv_columns(df1)
    c1 = find_city_col(df1)
    if c1: df1['City'] = df1[c1].astype(str).str.strip().str.title()
    date_col_f1 = next((c for c in df1.columns if 'date' in c.lower() or 'beat' in c.lower()), df1.columns[0])
    df1['DT_PARSED'] = pd.to_datetime(df1[date_col_f1], errors='coerce', dayfirst=True)
    df1['State'] = df1['State'].astype(str).str.strip().str.title() if 'State' in df1.columns else ''

    if not df2.empty:
        df2.columns = df2.columns.astype(str).str.strip()
        c2 = find_city_col(df2)
        if c2: df2['City'] = df2[c2].astype(str).str.strip().str.title()
        date_col_f2 = [c for c in df2.columns if 'DATE' in c.upper() or 'TIME' in c.upper()][0]
        df2['DT_PARSED'] = pd.to_datetime(df2[date_col_f2], errors='coerce', dayfirst=True)
        df2['State'] = df2['State'].astype(str).str.strip().str.title() if 'State' in df2.columns else ''
        df2['RANGE_CLEAN'] = df2['RANGE'].astype(str).str.lower().str.strip() if 'RANGE' in df2.columns else ''
        df2['is_within_30m'] = df2['RANGE_CLEAN'].apply(lambda x: 1 if 'within' in x and '30' in x else 0)
        df2['total_count'] = df2['RANGE_CLEAN'].apply(lambda x: 1 if 'within' in x or 'beyond' in x else 0)

    if not df3.empty:
        df3.columns = df3.columns.astype(str).str.strip()
        c3 = find_city_col(df3)
        if c3: df3['City'] = df3[c3].astype(str).str.strip().str.title()
        h_col3 = [c for c in df3.columns if 'hub' in c.lower()]
        if h_col3: df3['Hub'] = df3[h_col3[0]].astype(str).str.strip().str.upper()

        date_col_f3 = [c for c in df3.columns if 'DATE' in c.upper() or 'PICKUP' in c.upper() or 'TIME' in c.upper()][0]
        df3['DT_PARSED'] = pd.to_datetime(df3[date_col_f3], errors='coerce', dayfirst=True)
        df3['State'] = df3['State'].astype(str).str.strip().str.title() if 'State' in df3.columns else ''
        tat_cols = [col for col in df3.columns if 'TAT Tim' in col or 'TAT Timing' in col]
        df3['TAT_Minutes'] = df3[tat_cols[0]].apply(time_to_minutes) if tat_cols else 0
        df3['Radius_Clean'] = df3['Radius'].astype(str).str.lower().str.strip() if 'Radius' in df3.columns else 'overall'

    max_data_date = df1['DT_PARSED'].max()
    if pd.isna(max_data_date) and not df2.empty: max_data_date = df2['DT_PARSED'].max()

    dt_ftd = pd.to_datetime(ftd_date_raw, errors='coerce', dayfirst=True) if ftd_date_raw else max_data_date
    if pd.isna(dt_ftd): dt_ftd = max_data_date
    dt_mtd = pd.to_datetime(mtd_date_raw, errors='coerce', dayfirst=True) if mtd_date_raw else dt_ftd
    if pd.isna(dt_mtd): dt_mtd = dt_ftd
    mtd_start = dt_mtd.replace(day=1)

    df1_ftd = df1[df1['DT_PARSED'].dt.date == dt_ftd.date()].copy()
    df1_mtd = df1[(df1['DT_PARSED'].dt.date >= mtd_start.date()) & (df1['DT_PARSED'].dt.date <= dt_mtd.date())].copy()
    
    df2_ftd = df2[df2['DT_PARSED'].dt.date == dt_ftd.date()].copy() if not df2.empty else pd.DataFrame()
    df2_mtd = df2[(df2['DT_PARSED'].dt.date >= mtd_start.date()) & (df2['DT_PARSED'].dt.date <= dt_mtd.date())].copy() if not df2.empty else pd.DataFrame()
    
    df3_ftd = df3[df3['DT_PARSED'].dt.date == dt_ftd.date()].copy() if not df3.empty else pd.DataFrame()
    df3_mtd = df3[(df3['DT_PARSED'].dt.date >= mtd_start.date()) & (df3['DT_PARSED'].dt.date <= dt_mtd.date())].copy() if not df3.empty else pd.DataFrame()

    group_cols = ['State', 'City']
    ftd_overall = calculate_sheet_metrics(df1_ftd, df2_ftd, df3_ftd, group_cols, prefix="FTD ")
    mtd_overall = calculate_sheet_metrics(df1_mtd, df2_mtd, df3_mtd, group_cols, prefix="MTD ")
    overall_df = pd.merge(ftd_overall, mtd_overall, on=group_cols, how='outer')

    for pfx in ["FTD ", "MTD "]:
        if f'{pfx}Visit Compliance' in overall_df.columns: overall_df[f'{pfx}Visit Compliance'] = overall_df[f'{pfx}Visit Compliance'].fillna("-")
        if f'{pfx}On Time %' in overall_df.columns: overall_df[f'{pfx}On Time %'] = overall_df[f'{pfx}On Time %'].fillna("-")
        if f'{pfx}30mts %' in overall_df.columns: overall_df[f'{pfx}30mts %'] = overall_df[f'{pfx}30mts %'].fillna("0.0%")
        for tat_c in [f'{pfx}P2D-City', f'{pfx}P2D-Outskirt', f'{pfx}P2D-Overall']:
            if tat_c in overall_df.columns: overall_df[tat_c] = overall_df[tat_c].fillna("00:00")

    overall_df = overall_df.sort_values(by=['State', 'City'])

    overall_df = add_total_row_to_df(overall_df, df1_ftd, df2_ftd, df3_ftd, df1_mtd, df2_mtd, df3_mtd, first_col='State')

    chennai_df = process_hub_data(df1_ftd, df2_ftd, df3_ftd, df1_mtd, df2_mtd, df3_mtd, "Chennai")
    hitech_chennai_df = process_hub_data(df1_ftd, df2_ftd, df3_ftd, df1_mtd, df2_mtd, df3_mtd, "Hitech Chennai")

    return overall_df, chennai_df, hitech_chennai_df, dt_ftd.strftime('%d-%m-%Y'), dt_mtd.strftime('%d-%m-%Y'), df3

# ==========================================
# PROJECT 2: FIRST VISIT STATUS LOGIC (PASTEL MILD COLORS)
# ==========================================
def apply_excel_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    delay_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Mild Red
    ontime_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # Mild Green
    yet_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")    # Mild Yellow
    
    status_col_idx = next((col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value == 'Status'), None)
    if status_col_idx:
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=status_col_idx).value).strip()
            row_fill = delay_fill if val == 'Delay' else (yet_fill if val == 'Yet to done' else ontime_fill)
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = row_fill
    wb.save(file_path)

def process_first_visit(master_excel_path):
    df = read_any_excel_file(master_excel_path)
    
    col_lookup = {}
    for original_col in df.columns:
        clean_name = str(original_col).replace('[', '').replace(']', '').strip().lower()
        col_lookup[clean_name] = original_col

    def get_cell_val(row, candidate_names):
        for name in candidate_names:
            c_clean = name.replace('[', '').replace(']', '').strip().lower()
            if c_clean in col_lookup:
                actual_col = col_lookup[c_clean]
                val = row.get(actual_col)
                if pd.notna(val) and str(val).strip() not in ['-', '', 'nan', 'NaN', 'None']:
                    return val
        return None

    la_best_records = {}

    for idx, row in df.iterrows():
        mapped_la = get_cell_val(row, ['mapped la', 'mapped_la', 'la name', 'la'])
        if not mapped_la: continue
        mapped_la = str(mapped_la).strip()

        city_val = get_cell_val(row, ['city']) or ''
        city_val = str(city_val).strip()

        record_key = (city_val.lower(), mapped_la.lower())

        first_slot_index = None
        for i in range(1, 41):
            v_val = get_cell_val(row, [f'v{i}', f'v_{i}'])
            if v_val is not None:
                first_slot_index = i
                break
                
        if first_slot_index is not None:
            v_val_raw = get_cell_val(row, [f'v{first_slot_index}', f'v_{first_slot_index}'])
            v_time = parse_time_value(v_val_raw)
            
            if v_time is not None:
                v_mins = time_to_total_minutes(v_time)
                
                if record_key not in la_best_records or v_mins < la_best_records[record_key]['v_mins']:
                    p_val_raw = get_cell_val(row, [f'p{first_slot_index}', f'p_{first_slot_index}'])
                    p_time = parse_time_value(p_val_raw) if p_val_raw else None
                    p_mins = time_to_total_minutes(p_time) if p_time else None
                    
                    if p_val_raw is None or p_time is None:
                        status_value, tat_value = "Yet to done", "#####"
                    else:
                        time_diff = p_mins - v_mins if p_mins else None
                        if time_diff and time_diff > 0:
                            status_value, tat_value = "Delay", f"{time_diff // 60:02d}:{time_diff % 60:02d}"
                        else:
                            status_value, tat_value = "On time", "#####"

                    bu_value = get_cell_val(row, ['bu', 'd'])
                    bu_value = str(bu_value).strip() if bu_value else 'Unknown BU'
                    
                    client_val = get_cell_val(row, ['client name', 'client']) or ''
                    hub_val = get_cell_val(row, ['hub name', 'hub']) or ''
                    
                    la_name = get_cell_val(row, [f'la {first_slot_index}', f'la_{first_slot_index}'])
                    if not la_name: la_name = mapped_la

                    la_best_records[record_key] = {
                        'v_mins': v_mins,
                        'data': {
                            'BU': bu_value,
                            '[CLIENT NAME]': client_val,
                            '[HUB NAME]': hub_val,
                            'CITY': city_val,
                            '[MAPPED LA]': mapped_la,
                            'V1': v_time.strftime('%H:%M'),
                            'P1': p_time.strftime('%H:%M') if p_time else '-',
                            '[LA 1]': la_name,
                            'TAT': tat_value,
                            'Status': status_value
                        }
                    }

    result_df = pd.DataFrame([val['data'] for val in la_best_records.values()])
    return result_df.sort_values(by=['CITY', '[MAPPED LA]']) if not result_df.empty else result_df

# ==========================================
# PROJECT 6: LA VISIT STATUS LOGIC (EXCLUDING S1, S2... COLUMNS)
# ==========================================
def process_la_visit_status(file_path, output_path):
    df = read_any_excel_file(file_path)
    df.columns = df.columns.astype(str).str.strip()

    cols_to_keep = []
    for c in df.columns:
        clean_c = re.sub(r'[\[\]]', '', str(c)).strip().lower()
        if not re.match(r'^s\d+$', clean_c):
            cols_to_keep.append(c)

    df = df[cols_to_keep].copy()
    clean_col_map = {re.sub(r'[\[\]]', '', str(c)).strip().lower(): c for c in df.columns}

    for c in df.columns:
        clean_c = re.sub(r'[\[\]]', '', str(c)).strip().lower()
        if re.match(r'^[vp]\d+$', clean_c):
            df[c] = df[c].apply(format_hhmm_string)

    slot_numbers = []
    for c in df.columns:
        clean_c = re.sub(r'[\[\]]', '', str(c)).strip().lower()
        m = re.match(r'^la\s*(\d+)$', clean_c)
        if m:
            slot_numbers.append(m.group(1))

    slot_numbers = sorted(list(set(slot_numbers)), key=lambda x: int(x))

    new_columns = []
    for col in df.columns:
        new_columns.append(col)
        clean_c = re.sub(r'[\[\]]', '', str(col)).strip().lower()
        m = re.match(r'^la\s*(\d+)$', clean_c)
        if m:
            idx = m.group(1)
            new_columns.append(f"Diff{idx}")
            new_columns.append(f"Status{idx}")

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    calculated_data = {c: df[c].tolist() for c in df.columns}
    fill_map = {}

    for idx in slot_numbers:
        v_col_name = clean_col_map.get(f"v{idx}")
        p_col_name = clean_col_map.get(f"p{idx}")

        diff_list = []
        status_list = []

        for row_idx, row in df.iterrows():
            v_raw = row.get(v_col_name) if v_col_name else None
            p_raw = row.get(p_col_name) if p_col_name else None

            v_mins = time_to_total_minutes(parse_time_value(v_raw)) if v_raw else None
            p_mins = time_to_total_minutes(parse_time_value(p_raw)) if p_raw else None

            if v_mins is not None and p_mins is not None:
                diff_mins = p_mins - v_mins
                abs_diff = abs(diff_mins)
                diff_str = f"{abs_diff // 60:02d}:{abs_diff % 60:02d}"

                if -15 <= diff_mins <= 15:
                    st_val = "On time"
                    f_style = green_fill
                elif -30 <= diff_mins <= -16:
                    st_val = "Early"
                    f_style = amber_fill
                elif diff_mins < -30:
                    st_val = "Too early"
                    f_style = red_fill
                else:
                    st_val = "Delay"
                    f_style = red_fill

                diff_list.append(diff_str)
                status_list.append(st_val)
                fill_map[(row_idx, f"Diff{idx}")] = f_style
                fill_map[(row_idx, f"Status{idx}")] = f_style
            else:
                diff_list.append("-")
                status_list.append("-")

        calculated_data[f"Diff{idx}"] = diff_list
        calculated_data[f"Status{idx}"] = status_list

    out_df = pd.DataFrame()
    for col in new_columns:
        if col in calculated_data:
            out_df[col] = calculated_data[col]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LA Visit Status"

    ws.append(list(out_df.columns))

    for row_values in out_df.values:
        ws.append(list(row_values))

    col_name_to_idx = {name: i + 1 for i, name in enumerate(out_df.columns)}

    for (r_i, col_name), f_style in fill_map.items():
        c_idx = col_name_to_idx.get(col_name)
        if c_idx:
            ws.cell(row=r_i + 2, column=c_idx).fill = f_style

    wb.save(output_path)

# ==========================================
# PROJECT 3: BARCODE PENDING LOGIC & PDF
# ==========================================
def generate_barcode_pdf(missing_barcodes, pdf_path, start_bc, end_bc):
    doc = SimpleDocTemplate(
        pdf_path, 
        pagesize=letter, 
        rightMargin=15, 
        leftMargin=15, 
        topMargin=15, 
        bottomMargin=15
    )
    styles = getSampleStyleSheet()
    
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'],
        fontSize=9, leading=10, alignment=1, textColor=colors.HexColor('#000000')
    )
    
    elements = []

    if not missing_barcodes or "Invalid" in str(missing_barcodes[0]):
        title_style = ParagraphStyle(
            'TitleStyle', parent=styles['Heading1'],
            fontSize=14, leading=18, textColor=colors.HexColor('#1e3d59'), alignment=1
        )
        elements.append(Paragraph(f"<b>{missing_barcodes[0] if missing_barcodes else 'No barcodes found'}</b>", title_style))
    else:
        page_capacity = 120
        chunk_size = 30
        
        for page_idx in range(0, len(missing_barcodes), page_capacity):
            page_barcodes = missing_barcodes[page_idx : page_idx + page_capacity]
            col_chunks = [page_barcodes[i:i + chunk_size] for i in range(0, len(page_barcodes), chunk_size)]
            
            table_data = []
            for row_idx in range(chunk_size):
                row_cells = []
                for col_idx in range(4):
                    if col_idx < len(col_chunks) and row_idx < len(col_chunks[col_idx]):
                        val = str(col_chunks[col_idx][row_idx])
                    else:
                        val = ""
                    row_cells.append(Paragraph(val, cell_style))
                table_data.append(row_cells)
            
            t = Table(table_data, colWidths=[145, 145, 145, 145])
            t.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bbbbbb')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
            ]))
            
            elements.append(t)
            if page_idx + page_capacity < len(missing_barcodes):
                elements.append(PageBreak())

    doc.build(elements)

def process_barcode_pending(excel_file_path, start_bc, end_bc):
    df = read_any_excel_file(excel_file_path)
    df.columns = df.columns.astype(str).str.strip()
    bc_col = [c for c in df.columns if 'barcode' in c.lower()]
    if not bc_col:
        raise ValueError("Could not find 'Barcode' column in the uploaded Excel file.")
    
    existing_barcodes = set(df[bc_col[0]].dropna().astype(str).str.strip().str.upper())
    
    s_clean = start_bc.strip().upper()
    e_clean = end_bc.strip().upper()

    s_match = re.match(r"^([A-Za-z\-_]*)(0*\d+|\d+)$", s_clean)
    e_match = re.match(r"^([A-Za-z\-_]*)(0*\d+|\d+)$", e_clean)

    missing_barcodes = []
    
    if s_match and e_match:
        prefix_s, num_s_str = s_match.groups()
        prefix_e, num_e_str = e_match.groups()

        if prefix_s == prefix_e:
            prefix = prefix_s
            s_num = int(num_s_str)
            e_num = int(num_e_str)
            num_len = len(num_s_str)

            for num in range(s_num, e_num + 1):
                formatted_bc = f"{prefix}{num:0{num_len}d}"
                if formatted_bc not in existing_barcodes:
                    missing_barcodes.append(formatted_bc)
        else:
            missing_barcodes = ["Start and End Barcode Prefixes do not match! (e.g. both must start with CHL)"]
    else:
        missing_barcodes = ["Invalid Barcode Range format. Example valid format: CHL410001 to CHL425000"]

    return missing_barcodes

# ==========================================
# PROJECT 4: 3 IN 1 MERGE LOGIC
# ==========================================
def process_single_sheet_filter(file_path):
    df = read_any_excel_file(file_path)
    df.columns = df.columns.astype(str).str.strip()
    
    df = format_custom_dates(df)
    city_col = find_city_col(df)
    
    if city_col:
        df['city_clean'] = df[city_col].astype(str).str.strip().str.lower()
        df_filtered = df[df['city_clean'].isin(CITY_STATE_MAP.keys())].copy()
        df_filtered['State'] = df_filtered['city_clean'].map(CITY_STATE_MAP)
        df_filtered.drop(columns=['city_clean'], inplace=True)
        return df_filtered
    else:
        df['State'] = ''
        return df

def process_three_in_one(bv_path, ev_path, tat_path, output_excel_path):
    df_bv = process_single_sheet_filter(bv_path)
    df_ev = process_single_sheet_filter(ev_path)
    df_tat = process_single_sheet_filter(tat_path)

    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_bv.to_excel(writer, sheet_name='BV', index=False)
        df_ev.to_excel(writer, sheet_name='EV', index=False)
        df_tat.to_excel(writer, sheet_name='TAT', index=False)

# ==========================================
# PROJECT 5: DATA CAPTURE LOGIC
# ==========================================
def clean_str(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    return s.lower()

def process_data_capture(data_file_path, cluster_file_path, output_path):
    df_data = read_any_excel_file(data_file_path)
    df_cluster = read_any_excel_file(cluster_file_path)
    
    df_data.columns = df_data.columns.astype(str).str.strip()
    df_cluster.columns = df_cluster.columns.astype(str).str.strip()

    vid_col = next((c for c in df_data.columns if 'vidcapture' in c.lower().replace('_', '')), None)
    lab_col = next((c for c in df_data.columns if 'labdrop' in c.lower().replace(' ', '').replace('_', '')), None)
    status_col = next((c for c in df_data.columns if c.lower() == 'status'), None)
    
    loc_col = next((c for c in df_data.columns if 'loc' in c.lower() and 'name' in c.lower()), None)
    if not loc_col:
        loc_col = next((c for c in df_data.columns if 'loc' in c.lower() or 'client' in c.lower()), df_data.columns[0])

    client_col = next((c for c in df_cluster.columns if 'client' in c.lower() and 'name' in c.lower()), None)
    if not client_col:
        client_col = next((c for c in df_cluster.columns if 'client' in c.lower() or 'loc' in c.lower()), df_cluster.columns[0])

    micro_col = next((c for c in df_cluster.columns if 'micromarket' in c.lower().replace(' ', '').replace('_', '') or 'micro' in c.lower()), None)
    if not micro_col:
        micro_col = next((c for c in df_cluster.columns if 'cluster' in c.lower()), df_cluster.columns[-1])

    tat_results = []
    if vid_col and lab_col:
        s_vid = pd.to_datetime(df_data[vid_col], errors='coerce', dayfirst=True)
        s_lab = pd.to_datetime(df_data[lab_col], errors='coerce', dayfirst=True)
        diff_mins = (s_vid - s_lab).dt.total_seconds() / 60.0
        
        for idx, row in df_data.iterrows():
            curr_status = str(row[status_col]).strip() if status_col and pd.notna(row[status_col]) else "Done"
            if curr_status.lower() != 'done':
                tat_results.append(curr_status)
            else:
                m = diff_mins.iloc[idx]
                if pd.isna(m):
                    tat_results.append("-")
                elif m < 30:
                    tat_results.append("<30mins")
                else:
                    tat_results.append(">30mins")
    else:
        tat_results = ["-"] * len(df_data)

    df_data['TAT'] = tat_results

    micro_results = []
    if client_col and micro_col:
        micro_lookup = {}
        for _, c_row in df_cluster.iterrows():
            raw_c_name = clean_str(c_row[client_col])
            raw_m_val = str(c_row[micro_col]).strip() if pd.notna(c_row[micro_col]) else ""
            if raw_c_name and raw_m_val and raw_m_val.lower() != 'nan':
                micro_lookup[raw_c_name] = raw_m_val

        for _, d_row in df_data.iterrows():
            raw_l_name = clean_str(d_row[loc_col])
            micro_results.append(micro_lookup.get(raw_l_name, "-"))
    else:
        micro_results = ["-"] * len(df_data)

    df_data['Micromarket'] = micro_results
    df_data.to_excel(output_path, index=False)

# ==========================================
# ROUTES & CONTROLLERS
# ==========================================
@app.route('/')
def home():
    return render_template('home.html')

@app.route('/beat-performance', methods=['GET', 'POST'])
def beat_performance():
    if request.method == 'POST':
        try:
            ftd_date, mtd_date = request.form.get('ftd_date'), request.form.get('mtd_date')
            master_file = request.files['master_file']
            file_path = os.path.join(UPLOAD_FOLDER, f"{int(time.time())}_{master_file.filename}")
            master_file.save(file_path)
            
            overall, chennai, hitech, disp_ftd, disp_mtd, df3_raw = process_beat_performance(file_path, ftd_date, mtd_date)
            
            output_path = os.path.join(UPLOAD_FOLDER, "beat_output.xlsx")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                overall.to_excel(writer, sheet_name='Overall', index=False)
                if not chennai.empty: chennai.to_excel(writer, sheet_name='Chennai', index=False)
                if not hitech.empty: hitech.to_excel(writer, sheet_name='Hitech Chennai', index=False)
                write_not_dropped_sheet(writer, df3_raw)
            
            return render_template('beat_performance.html', data=overall.to_dict(orient='records'), columns=overall.columns.tolist(), ftd_date=disp_ftd, mtd_date=disp_mtd, show_table=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
    return render_template('beat_performance.html', show_table=False)

@app.route('/beat-summary', methods=['GET', 'POST'])
def beat_summary():
    if request.method == 'POST':
        try:
            master_file = request.files.get('master_file')
            from_date = request.form.get('from_date')
            to_date = request.form.get('to_date')
            selected_state = request.form.get('selected_state', 'All')
            selected_city = request.form.get('selected_city', 'All')
            compare_city = request.form.get('compare_city', 'None')

            if master_file:
                file_path = os.path.join(UPLOAD_FOLDER, f"sum_{int(time.time())}_{master_file.filename}")
                master_file.save(file_path)
            else:
                return "<h3>Please upload master Excel dump</h3>"

            overall, _, _, _, _, _ = process_beat_performance(file_path, from_date, to_date)

            states_list = [s for s in overall['State'].unique() if s and s != 'Total/Avg']
            cities_list = [c for c in overall['City'].unique() if c and c != 'Total/Avg']

            filtered_df = overall.copy()

            if selected_state != 'All':
                filtered_df = filtered_df[filtered_df['State'] == selected_state]

            if compare_city != 'None' and compare_city in cities_list:
                filtered_df = filtered_df[filtered_df['City'].isin([selected_city, compare_city])]
            elif selected_city != 'All':
                filtered_df = filtered_df[filtered_df['City'] == selected_city]

            return render_template('beat_summary.html', 
                                   data=filtered_df.to_dict(orient='records'), 
                                   columns=filtered_df.columns.tolist(), 
                                   states_list=states_list, 
                                   cities_list=cities_list, 
                                   selected_state=selected_state, 
                                   selected_city=selected_city, 
                                   compare_city=compare_city, 
                                   from_date=from_date, 
                                   to_date=to_date, 
                                   show_dropdowns=True, 
                                   show_table=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
    return render_template('beat_summary.html', show_dropdowns=False, show_table=False)

@app.route('/la-visit-status', methods=['GET', 'POST'])
def la_visit_status():
    if request.method == 'POST':
        try:
            bv_file = request.files['bv_file']
            file_path = os.path.join(UPLOAD_FOLDER, f"la_v_{int(time.time())}_{bv_file.filename}")
            bv_file.save(file_path)
            
            output_path = os.path.join(UPLOAD_FOLDER, "la_visit_status_output.xlsx")
            process_la_visit_status(file_path, output_path)
            
            return render_template('la_visit_status.html', show_result=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
    return render_template('la_visit_status.html', show_result=False)

@app.route('/first-visit', methods=['GET', 'POST'])
def first_visit():
    if request.method == 'POST':
        try:
            master_file = request.files['master_file']
            file_path = os.path.join(UPLOAD_FOLDER, f"{int(time.time())}_{master_file.filename}")
            master_file.save(file_path)
            
            result_df = process_first_visit(file_path)
            output_path = os.path.join(UPLOAD_FOLDER, "first_visit_output.xlsx")
            result_df.to_excel(output_path, index=False)
            apply_excel_formatting(output_path)
            
            return render_template('first_visit.html', data=result_df.to_dict(orient='records'), columns=result_df.columns.tolist(), show_table=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
    return render_template('first_visit.html', show_table=False)

@app.route('/barcode-pending', methods=['GET', 'POST'])
def barcode_pending():
    if request.method == 'POST':
        try:
            start_bc = request.form.get('start_bc')
            end_bc = request.form.get('end_bc')
            master_file = request.files['master_file']
            
            file_path = os.path.join(UPLOAD_FOLDER, f"{int(time.time())}_{master_file.filename}")
            master_file.save(file_path)
            
            missing_barcodes = process_barcode_pending(file_path, start_bc, end_bc)
            
            pdf_path = os.path.join(UPLOAD_FOLDER, "barcode_pending.pdf")
            generate_barcode_pdf(missing_barcodes, pdf_path, start_bc, end_bc)
            
            return render_template('barcode_pending.html', missing_barcodes=missing_barcodes, total_missing=len(missing_barcodes), start_bc=start_bc, end_bc=end_bc, show_result=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
    return render_template('barcode_pending.html', show_result=False)

@app.route('/three-in-one', methods=['GET', 'POST'])
def three_in_one():
    if request.method == 'POST':
        try:
            bv_file = request.files['bv_file']
            ev_file = request.files['ev_file']
            tat_file = request.files['tat_file']
            
            bv_path = os.path.join(UPLOAD_FOLDER, f"bv_{int(time.time())}_{bv_file.filename}")
            ev_path = os.path.join(UPLOAD_FOLDER, f"ev_{int(time.time())}_{ev_file.filename}")
            tat_path = os.path.join(UPLOAD_FOLDER, f"tat_{int(time.time())}_{tat_file.filename}")
            
            bv_file.save(bv_path)
            ev_file.save(ev_path)
            tat_file.save(tat_path)
            
            merged_output_path = os.path.join(UPLOAD_FOLDER, "merged_master.xlsx")
            process_three_in_one(bv_path, ev_path, tat_path, merged_output_path)
            
            return render_template('three_in_one.html', show_result=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
            
    return render_template('three_in_one.html', show_result=False)

@app.route('/data-capture', methods=['GET', 'POST'])
def data_capture():
    if request.method == 'POST':
        try:
            data_file = request.files['data_file']
            cluster_file = request.files['cluster_file']
            
            data_path = os.path.join(UPLOAD_FOLDER, f"data_{int(time.time())}_{data_file.filename}")
            cluster_path = os.path.join(UPLOAD_FOLDER, f"cluster_{int(time.time())}_{cluster_file.filename}")
            
            data_file.save(data_path)
            cluster_file.save(cluster_path)
            
            dc_output_path = os.path.join(UPLOAD_FOLDER, "data_capture_output.xlsx")
            process_data_capture(data_path, cluster_path, dc_output_path)
            
            return render_template('data_capture.html', show_result=True)
        except Exception as e:
            return f"<h3>Error: {e}</h3>"
            
    return render_template('data_capture.html', show_result=False)

@app.route('/download/<report_type>')
def download(report_type):
    if report_type == 'beat': file_name = "beat_output.xlsx"
    elif report_type == 'firstvisit': file_name = "first_visit_output.xlsx"
    elif report_type == 'barcode_pdf': file_name = "barcode_pending.pdf"
    elif report_type == 'merged_excel': file_name = "merged_master.xlsx"
    elif report_type == 'data_capture': file_name = "data_capture_output.xlsx"
    elif report_type == 'la_visit_status': file_name = "la_visit_status_output.xlsx"
    return send_file(os.path.join(UPLOAD_FOLDER, file_name), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
